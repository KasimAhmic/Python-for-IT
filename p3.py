import openpyxl
import sys
from twilio.rest import Client

# Analyzes the spreadsheet and returns a message with missing grades
def get_missing_grades(sheet):
    missing_grades = 0
    students = []

    # Iterate through each row and column of the spreadsheet
    for i in range(sheet.max_row):
        for x in range(sheet.max_column):
            # If the current cell is empty
            if not sheet.cell(row=i + 1, column=x + 1).value:
                # Increment the missing_grades counter by one
                missing_grades += 1
                name = sheet.cell(row=i + 1, column=1).value
                
                # If the student isn't already in the students list, add them
                if name not in students:
                    students.append(name)

    # If there are more than 0 grades
    if missing_grades > 0:
        # Create a message specifying how many missing grades there are
        message = 'you have ' + str(missing_grades) + ' missing grades, for the following students:'

        # Append all the students who have missing grades to the message created earlier
        for student in students:
            message += '\n' + student

        return message
    # If there are no missing grades
    else:
        # Return a message saying there are no missing grades
        return 'you have no students with missing grades.'

# Sends a text of the missing grades
def send_text(message, sid, token, to_num, from_num):
    # Create Client object and send text using provided information
    client = Client(sid, token)
    text = client.messages.create(body='Our records show that ' + message, from_=from_num, to=to_num)

def main():
    # Specify file name
    file_name = sys.argv[1]

    # If the user specifies Twilio information, create variables for it
    if (len(sys.argv) > 2):
        twilio_sid = sys.argv[2]
        auth_token = sys.argv[3]
        to_number = sys.argv[4]
        from_number = sys.argv[5]

    # Try to open the file
    try:
        workbook = openpyxl.load_workbook(file_name)
    # If the file fails to open
    except FileNotFoundError:
        # Print an error message
        print('Error: Invalid filename given.')
        return

    # Get the sheet from the workbook
    sheet = workbook['Sheet1']

    # Run the get_missing_grades function and assign the returned message to the message variable
    message = get_missing_grades(sheet)
    print(message[0].capitalize() + message[1:])

    # If the user specified Twilio information
    if (len(sys.argv) > 2):
        # Try to send a text
        try:
            send_text(message, twilio_sid, auth_token, to_number, from_number)
        # If the text fails, prompt user to check the Twilio information
        except:
            print('\nText could not be sent. Please verify To (' + to_number + ') and From (' + from_number + ') numbers as well as your SID and Auth Token.')
            return
        
        print()
        print('The following message will be sent to ' + to_number + ':')
        print('Our records show that ' + message)

if __name__ == '__main__':
    main()